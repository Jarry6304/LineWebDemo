#!/usr/bin/env python3
"""
從 data/*.csv 產生 data/site.xlsx 模板。

用途：給維護者一份「上傳到 Google Sheets 的初始模板」。
之後維護者只在 Google Sheets 上編輯，這個 xlsx 不需要再改。
有需要重新生成時，執行：

    python3 scripts/build_site_xlsx.py

需求：openpyxl
    pip install openpyxl
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
OUT_PATH = DATA_DIR / "site.xlsx"

# ============================================================
# Sheet 規範：順序、欄位、寬度、說明
# 程式從上到下讀取，每張表第 1 列為欄位名（程式依賴）
# ============================================================
SHEETS = [
    {
        "name": "config",
        "csv": "config.csv",
        "fields": [
            ("key",   "設定鍵（請勿修改）", 28),
            ("value", "設定值",             60),
            ("note",  "說明（給維護者看的）", 40),
        ],
    },
    {
        "name": "classes",
        "csv": "classes.csv",
        "fields": [
            ("class_id",                "班級代號（英數小寫，不可重複）",          18),
            ("name",                    "班級名稱（例：山柿班）",                  14),
            ("series",                  "所屬系列（須與 series.name 完全一致）",   14),
            ("level",                   "難度（初階/中階/進階/挑戰/親子/探究）",   12),
            ("region",                  "地區",                                     10),
            ("price",                   "費用（整數）",                             10),
            ("age_range",               "年齡層",                                   14),
            ("experience_requirement",  "經驗代號（須對應 exp_hints.code）",       14),
            ("duration_hours",          "每堂時數（整數）",                         10),
            ("transport_note",          "交通說明",                                 22),
            ("activity_areas",          "活動範圍",                                 30),
            ("description",             "課程簡介（顯示於卡片）",                   40),
            ("display_order",           "顯示順序（小到大，整數）",                 10),
            ("price_unit",              "價格單位（例：／每位小孩；無則留空）",     16),
            ("fee_included",            "費用包含項目（詳情頁摺疊；留空則不顯示）", 40),
        ],
    },
    {
        "name": "course_dates",
        "csv": "course_dates.csv",
        "fields": [
            ("class_id",     "班級代號（須與 classes.class_id 一致）",  18),
            ("session_no",   "第幾堂（整數）",                          10),
            ("date",         "日期（YYYY-MM-DD，純文字）",              14),
            ("weekday",      "週幾（一字：六、日 …）",                  8),
            ("time_period",  "時段（full_day / morning / afternoon）",  12),
            ("notes",        "備註（可空）",                            24),
        ],
    },
    {
        "name": "faqs",
        "csv": "faqs.csv",
        "fields": [
            ("order",     "顯示順序（整數，小到大）", 8),
            ("category",  "分類（退費／報名／安全 …）", 12),
            ("question",  "問題",                  40),
            ("answer",    "答案（可 Alt+Enter 換行）", 60),
        ],
    },
    {
        "name": "series",
        "csv": "series.csv",
        "fields": [
            ("series_id",     "系列代號（英數小寫，當網址錨點，例：mountain）", 14),
            ("name",          "系列名稱（須與 classes.series 一致）", 14),
            ("title",         "區塊標題（例：山野成長系列）", 18),
            ("tagline",       "課程頁分區副標", 36),
            ("preview_desc",  "首頁系列卡片簡介", 40),
            ("display_order", "顯示順序", 10),
        ],
    },
    {
        "name": "keywords",
        "csv": "keywords.csv",
        "fields": [
            ("display_order", "顯示順序", 10),
            ("name",          "關鍵字（例：自然觀察）", 14),
            ("description",   "說明（點擊展開時顯示）", 50),
        ],
    },
    {
        "name": "about",
        "csv": "about.csv",
        "fields": [
            ("display_order", "顯示順序", 10),
            ("style",         "樣式：lead/muted（常顯）、body/emphasis（收進「閱讀完整故事」）", 14),
            ("text",          "段落文字（可 Alt+Enter 換行）", 60),
        ],
    },
    {
        "name": "refund",
        "csv": "refund.csv",
        "fields": [
            ("display_order", "顯示順序", 10),
            ("stage",         "退費階段（例：開課日前 21 天）", 24),
            ("ratio",         "退費比例（例：扣除行政費 $200 後全額退費）", 40),
        ],
    },
    {
        "name": "privacy",
        "csv": "privacy.csv",
        "fields": [
            ("display_order", "顯示順序", 10),
            ("heading",       "小節標題（例：蒐集目的）", 18),
            ("body",          "段落內容（可 Alt+Enter 換行）", 60),
        ],
    },
    {
        "name": "exp_hints",
        "csv": "exp_hints.csv",
        "fields": [
            ("code", "經驗代號（須與 classes.experience_requirement 一致）", 14),
            ("hint", "顯示文字（例：低頻率戶外或爬山經驗）", 40),
        ],
    },
]

HEADER_FILL = PatternFill("solid", fgColor="2D5F3F")
HEADER_FONT = Font(name="Noto Sans TC", color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
BODY_FONT = Font(name="Noto Sans TC", size=11)
THIN = Side(border_style="thin", color="D6CFC0")
CELL_BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def write_intro_sheet(wb):
    ws = wb.create_sheet("說明", 0)
    ws.sheet_properties.tabColor = "A4C763"

    intro = [
        ("虎甲自然網站　資料規範", True, 16),
        ("", False, 11),
        ("這份試算表是網站所有內容的單一來源。維護者只要在這裡編輯，", False, 11),
        ("網站約 5 分鐘內就會自動更新（透過 Google Sheets 線上 CSV）。", False, 11),
        ("", False, 11),
        ("【sheet 一覽】", True, 12),
        ("　config        ── 站台設定 + 全站文字（key/value/note）", False, 11),
        ("　classes       ── 班級清單（一列一班）", False, 11),
        ("　course_dates  ── 課程日期（一列一堂，用 class_id 關聯班級）", False, 11),
        ("　faqs          ── 常見問題（一列一題）", False, 11),
        ("　series        ── 三大系列（首頁卡片 + 課程頁分區）", False, 11),
        ("　keywords      ── 首頁四宮格", False, 11),
        ("　about         ── 首頁品牌故事段落", False, 11),
        ("　refund        ── 退費規定表", False, 11),
        ("　privacy       ── 個資聲明段落", False, 11),
        ("　exp_hints     ── 經驗代號對照", False, 11),
        ("", False, 11),
        ("【規格驅動原則】", True, 12),
        ("　HTML 不寫任何文案，全部文字來自這份表。", False, 11),
        ("　程式只負責渲染，不寫任何商業邏輯（系列分區、價格單位等都看欄位）。", False, 11),
        ("　漏鍵不會壞掉：HTML 內保留一份 fallback 原文。", False, 11),
        ("", False, 11),
        ("【新增班級的步驟】", True, 12),
        ("　1. classes 新增一列，填好 15 欄（class_id 不可重複；series 須對得到 series.name）", False, 11),
        ("　2. course_dates 新增該班的每堂課（class_id 要一致）", False, 11),
        ("　3. 存檔即可，網站約 5 分鐘內自動更新", False, 11),
        ("", False, 11),
        ("【新增系列】", True, 12),
        ("　series 新增一列（取好 series_id），首頁與課程頁都會自動長出。", False, 11),
        ("", False, 11),
        ("【新增 FAQ／退費規定／個資段／關鍵字／故事段落】", True, 12),
        ("　到對應 sheet 新增一列即可。display_order 設大一點就會排到最後。", False, 11),
        ("", False, 11),
        ("【修改站台設定／文字】", True, 12),
        ("　到 config 表改 value 欄；key 欄勿動。", False, 11),
        ("　多行文字（home_hero_title 等）用 \\n 代表換行。", False, 11),
        ("", False, 11),
        ("【⚠️ 資安守則 — 開動前務必讀】", True, 14),
        ("　這份試算表上傳到 Google Drive 後會被全網路讀取。", False, 11),
        ("　網站把試算表 ID 寫在 JS 裡（任何訪客都能看到），所以 ID = 公開的入口。", False, 11),
        ("", False, 11),
        ("　🔴 絕對不要放進這份試算表：", True, 11),
        ("　　• 學員 / 家長個資（姓名、電話、身分證、緊急聯絡）", False, 11),
        ("　　• 信用卡、銀行帳號、轉帳記錄", False, 11),
        ("　　• 內部成本、毛利、講師時薪", False, 11),
        ("　　• 講師私人事由、未公開的取消預定", False, 11),
        ("　　• API token、密碼、LINE Channel Token", False, 11),
        ("", False, 11),
        ("　⚠️ 重要：在同一份試算表新增一個「內部用」sheet 不會保密。", False, 11),
        ("　　即使網站不去讀，任何人用以下 URL 都能直接抓走：", False, 11),
        ("　　https://docs.google.com/spreadsheets/d/{ID}/gviz/tq?tqx=out:csv&sheet={名稱}", False, 10),
        ("", False, 11),
        ("　✅ 內部資料請用「另一份完全分開」的試算表：", True, 11),
        ("　　• 在 Google Drive 另建獨立檔案", False, 11),
        ("　　• 分享 = 「指定 email 邀請」，不可用「擁有連結者」", False, 11),
        ("　　• 該 ID 永遠不寫進 site-config.js 或任何網站檔案", False, 11),
        ("", False, 11),
        ("　新增資料前的 3 題自我檢查：", True, 11),
        ("　　1. 被截圖貼到 FB 社團我會困擾嗎？會 → 不要放這裡", False, 11),
        ("　　2. 涉及第三方（學員、講師、合作場地）嗎？涉及 → 不要放這裡", False, 11),
        ("　　3. 必須讓全網看到嗎？不必 → 不要放這裡", False, 11),
        ("", False, 11),
        ("【常見地雷】", True, 12),
        ("　• class_id、series_id 必須英數小寫，中文會壞", False, 11),
        ("　• 跨表關聯要一致：classes.series ↔ series.name、course_dates.class_id ↔ classes.class_id", False, 11),
        ("　• 日期欄保持 YYYY-MM-DD 純文字（選欄→格式→數字→純文字）", False, 11),
        ("　• 各表第 1 列欄位名不可改、不可刪欄插欄", False, 11),
        ("　• 新增 sheet 沒關係（程式只讀指定 10 張），但仍會被公網讀取（見上方資安守則）", False, 11),
    ]

    for idx, (text, bold, size) in enumerate(intro, start=1):
        cell = ws.cell(row=idx, column=1, value=text)
        cell.font = Font(name="Noto Sans TC", size=size, bold=bold,
                         color="2D5F3F" if bold else "333333")
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 100
    ws.sheet_view.showGridLines = False


def write_data_sheet(wb, spec, csv_path):
    ws = wb.create_sheet(spec["name"])
    ws.sheet_properties.tabColor = "F2C744"

    field_names = [f[0] for f in spec["fields"]]
    field_widths = [f[2] for f in spec["fields"]]

    for col_idx, (name, note, _w) in enumerate(spec["fields"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER
        cell.comment = Comment(note, "規範")

    for col_idx, w in enumerate(field_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 28

    rows = []
    if csv_path.exists():
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                rows.append([r.get(name, "") for name in field_names])

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = BODY_ALIGN
            cell.font = BODY_FONT
            cell.border = CELL_BORDER

    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 110


def main():
    wb = Workbook()
    wb.remove(wb.active)

    write_intro_sheet(wb)
    for spec in SHEETS:
        csv_path = DATA_DIR / spec["csv"]
        write_data_sheet(wb, spec, csv_path)

    wb.save(OUT_PATH)
    print(f"產生：{OUT_PATH.relative_to(DATA_DIR.parent)}")
    print(f"      sheet 數量：{len(wb.sheetnames)} → {wb.sheetnames}")


if __name__ == "__main__":
    main()
